"""
PDF to XLSX Converter – Streamlit App
"""

import io, re
import streamlit as st
import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

EXCLUDED_COLUMNS = {"threat score", "data"}
NUMERIC_COLUMNS  = {"threats", "blocked requests", "requests"}

# ─── Excel Styling ───────────────────────────────────────────────────
H_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
H_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
H_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
C_FONT = Font(name="Calibri", size=11)
C_ALIGN = Alignment(vertical="top", wrap_text=True)
N_ALIGN = Alignment(horizontal="right", vertical="top")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"), right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"), bottom=Side(style="thin", color="CCCCCC"),
)

def _san(n): return re.sub(r'[\\/*?:\[\]]', '', n)[:31]

# Strip emoji and illegal XML chars that Excel/openpyxl can't handle
_ILLEGAL_XML = re.compile(
    u'[^\u0009\u000A\u000D\u0020-\uD7FF\uE000-\uFFFD\u10000-\u10FFFF]')
def _strip_unicode(v):
    """Remove emoji and illegal XML characters from a string value."""
    if not isinstance(v, str): return v
    # Remove emoji (most fall in these blocks)
    v = re.sub(u'[\U0001F000-\U0001FFFF\U00002600-\U000027BF\U0001F300-\U0001F9FF]', '', v)
    # Remove remaining illegal XML characters
    v = _ILLEGAL_XML.sub('', v)
    return v.strip()

def _aw(ws):
    for cc in ws.columns:
        ml = 0; cl = get_column_letter(cc[0].column)
        for c in cc:
            try:
                l = max(len(s) for s in str(c.value or "").split("\n"))
                if l > ml: ml = l
            except: pass
        ws.column_dimensions[cl].width = min(ml + 4, 60)

def _sh(ws, r, n):
    for c in range(1, n+1):
        cell = ws.cell(row=r, column=c)
        cell.font = H_FONT; cell.fill = H_FILL
        cell.alignment = H_ALIGN; cell.border = BORDER

def _sc(cell, num=False):
    cell.font = C_FONT; cell.border = BORDER
    cell.alignment = N_ALIGN if num else C_ALIGN
    if num: cell.number_format = '#,##0'

# ─── Numeric conversion ─────────────────────────────────────────────
_SFX = {'': 1, 'k': 1_000, 'm': 1_000_000, 'b': 1_000_000_000, 't': 1_000_000_000_000}
_NR = re.compile(r'^(?P<num>[\d.,]+)\s*(?P<sfx>[KkMmBbTt]?)$')

def _cn(value):
    if value is None: return ""
    v = str(value).strip()
    if not v: return ""
    m = _NR.match(v)
    if not m: return value
    ns, sfx = m.group("num"), m.group("sfx").lower()
    if sfx and ',' in ns:
        p = ns.split(',')
        ns = p[0] + '.' + p[1] if len(p) == 2 and len(p[1]) <= 2 else ns.replace(',', '')
    else: ns = ns.replace(',', '')
    try: base = float(ns)
    except ValueError: return value
    r = base * _SFX.get(sfx, 1)
    return int(r) if r == int(r) else r

def _in(v): return isinstance(v, (int, float))

def _efn(value):
    if value is None: return ""
    v = str(value).strip()
    r = _cn(v)
    if _in(r): return r
    for tok in v.split():
        r = _cn(tok)
        if _in(r): return r
    return value

# ─── Word-coordinate helpers ─────────────────────────────────────────
def _gl(words, ytol=4):
    if not words: return []
    sw = sorted(words, key=lambda w: (w["top"], w["x0"]))
    out, cur = [], [sw[0]]
    for w in sw[1:]:
        if abs(w["top"] - cur[0]["top"]) <= ytol: cur.append(w)
        else: out.append(sorted(cur, key=lambda w: w["x0"])); cur = [w]
    if cur: out.append(sorted(cur, key=lambda w: w["x0"]))
    return out

def _lt(line): return " ".join(w["text"] for w in line).strip()

def _ih(text):
    t = text.lower()
    kws = ["threats", "threat url", "country", "signature id",
           "threat main type", "threat sub type", "threat score",
           "threat cve", "blocked requests", "requests", "domain", "name"]
    return sum(1 for kw in kws if kw in t) >= 2

def _is(text):
    return bool(re.match(
        r'(Top\s+Threats?\s+By\s+|Applications?\s+Summary|ZoneWAF)',
        text, re.IGNORECASE))

def _pc(lw):
    if not lw: return []
    sw = sorted(lw, key=lambda w: w["x0"])
    if len(sw) == 1:
        return [{"name": sw[0]["text"], "x0": sw[0]["x0"], "x1": sw[0]["x1"]}]
    gaps = [sw[i]["x0"] - sw[i-1]["x1"] for i in range(1, len(sw))]
    gv = sorted(gaps); bj, bi = 0, -1
    for i in range(len(gv)-1):
        j = gv[i+1] - gv[i]
        if j > bj: bj, bi = j, i
    thr = (gv[bi]+gv[bi+1])/2 if bj > 5 and bi >= 0 else max(gv[-1]+10, 20)
    cols, cur = [], [sw[0]]
    for i, g in enumerate(gaps):
        if g > thr:
            cols.append({"name": " ".join(w["text"] for w in cur),
                         "x0": cur[0]["x0"], "x1": cur[-1]["x1"]})
            cur = [sw[i+1]]
        else: cur.append(sw[i+1])
    if cur:
        cols.append({"name": " ".join(w["text"] for w in cur),
                     "x0": cur[0]["x0"], "x1": cur[-1]["x1"]})
    return cols

def _cb(hdrs):
    b = []
    for i, c in enumerate(hdrs):
        xs = 0 if i == 0 else (hdrs[i-1]["x1"]+c["x0"])/2
        xe = float("inf") if i == len(hdrs)-1 else (c["x1"]+hdrs[i+1]["x0"])/2
        b.append((xs, xe))
    return b

def _as(lw, hdrs):
    bounds = _cb(hdrs); row = [""]*len(hdrs)
    for w in sorted(lw, key=lambda w: w["x0"]):
        cx = (w["x0"]+w["x1"])/2
        for i, (xs, xe) in enumerate(bounds):
            if xs <= cx < xe: row[i] = (row[i]+" "+w["text"]).strip(); break
    return row

def _pp(row, hdrs):
    for i, h in enumerate(hdrs):
        if h["name"].lower().strip() not in NUMERIC_COLUMNS: continue
        cell = row[i].strip()
        if not cell: continue
        toks = cell.split(); nums, txts = [], []
        for t in toks:
            if _NR.match(t.replace(',', '')): nums.append(t)
            else: txts.append(t)
        if txts:
            for j in range(i-1, -1, -1):
                if hdrs[j]["name"].lower().strip() not in NUMERIC_COLUMNS:
                    prev = row[j].strip(); add = " ".join(txts)
                    row[j] = (prev+", "+add).strip(", ") if prev else add; break
            row[i] = " ".join(nums)
    return row

# ─── Text-based widget/card extraction (for non-tabular PDFs) ────────
def _extract_widgets_from_text(page_text):
    """Parse plain text from extract_text() to find metric cards and Top-N lists."""
    lists, metrics = {}, {}
    if not page_text: return lists, metrics
    
    raw_lines = [ln.strip() for ln in page_text.split('\n') if ln.strip()]
    
    # ── Ranked list detection ──
    # Titles like "Top 5 Sites Attacked", "Top 1 Countries by DDoS Attack Volume"
    list_title_re = re.compile(
        r'Top\s+\d+\s+'
        r'(Countries|Sites|Attack\s+Categories|Targeted\s+Applications|IPs)'
        r'.*', re.IGNORECASE)
    # Ranked items: "1 app.trustbancgroup.com 1.1M" or just "app.trustbancgroup.com 1.1M"
    ranked_re = re.compile(
        r'^(\d+)\s+(.+?)\s+([\d.,]+\s*[KkMmBbTt]?)$')
    # Also try without leading rank number (some PDFs skip the circled digit)
    pair_re = re.compile(
        r'^(.+?)\s{2,}([\d.,]+\s*[KkMmBbTt]?)$')
    
    cur_list = None
    cur_rank = 0
    
    for i, ln in enumerate(raw_lines):
        # Check for list title
        tm = list_title_re.search(ln)
        if tm:
            cur_list = ln.strip()
            if cur_list not in lists: lists[cur_list] = []
            cur_rank = 0
            continue
        
        # If we're inside a list, try to grab ranked items
        if cur_list is not None:
            # Try ranked format: "1 United Kingdom 1.1M"
            rm = ranked_re.match(ln)
            if rm:
                lists[cur_list].append([rm.group(1), rm.group(2).strip(), rm.group(3).strip()])
                continue
            # Try pair format: "United Kingdom       1.1M"
            pm = pair_re.match(ln)
            if pm:
                name_part = pm.group(1).strip()
                val_part = pm.group(2).strip()
                # Skip if looks like a section header
                if not list_title_re.search(name_part) and not re.match(r'^(Threat|Security|Bot|DDoS|Network|Platform|Compliance)', name_part, re.IGNORECASE):
                    cur_rank += 1
                    lists[cur_list].append([str(cur_rank), name_part, val_part])
                    continue
            # End the current list if we hit a new section or empty-ish line
            if list_title_re.search(ln) or re.match(r'^(Threat|Security|Bot|DDoS|Network|Platform|Compliance|Compared)', ln, re.IGNORECASE):
                cur_list = None
    
    # ── Inline metric detection ──
    # Patterns like "Total Attacks Blocked 1.1M 224586.81%"
    # or "SwyftComply Requested 0"
    inline_metric_re = re.compile(
        r'^(Total\s+Attacks\s+Blocked|Attacks\s+Blocked\s+by\s+Custom\s+Rules|'
        r'Total\s+Vulnerabilities\s+Detected|WAF\s+Rules\s+Created|'
        r'Custom\s+Rules\s+Created|Total\s+Bandwidth\s+Used|'
        r'SwyftComply\s+Requested|Scans\s+Completed|'
        r'Manual\s+Pentesting\s+Requested|POC\s+Requested|'
        r'Total\s+DDoS\s+Attacks\s+Blocked|Sites\s+under\s+DDoS\s+Attack|'
        r'Sites\s+under\s+Bot\s+Attack|Bot\s+Attacks\s+Blocked)'
        r'\s+([\d.,]+\s*[KkMmBbTtGgBb]*)\s*(.*)',
        re.IGNORECASE)
    
    for i, ln in enumerate(raw_lines):
        ll = ln.lower()
        
        # Inline metric on the same line
        im = inline_metric_re.match(ln)
        if im:
            metrics[im.group(1).strip()] = [im.group(2).strip(), im.group(3).strip()]
            continue
        
        # Block-style: label on one line, value on next line
        # e.g.  "Sites under Bot Attack"  then  "0"  on the next line
        block_labels = [
            "Sites under Bot Attack", "Bot Attacks Blocked",
            "Sites under DDoS Attack", "Total DDoS Attacks Blocked",
        ]
        for label in block_labels:
            if label.lower() in ll and len(ln) < len(label) + 15:
                # Look in the next few lines for a numeric value
                for j in range(i+1, min(i+4, len(raw_lines))):
                    nt = raw_lines[j].strip()
                    if re.match(r'^[\d.,]+\s*[KkMmBbTt]?$', nt):
                        if label not in metrics:
                            metrics[label] = [nt, ""]
                        break
                break
        
        # Detect "No Targeted Applications" etc.
        if "no targeted applications" in ll:
            metrics["Targeted Applications"] = ["0", "None detected"]
        if "no bot categories detected" in ll:
            metrics["Bot Categories"] = ["0", "None detected"]
    
    return lists, metrics

def _et(page, carry=None):
    words = page.extract_words(x_tolerance=3, y_tolerance=3, keep_blank_chars=False)
    if not words: return [], None
    lines = _gl(words); tables, cur_h, cur_r = [], carry, []
    ic = carry is not None; hy = -1 if carry else None
    for line in lines:
        lt = _lt(line)
        if not lt: continue
        if _ih(lt):
            if cur_h and cur_r: tables.append({"h": cur_h, "r": cur_r, "c": ic})
            cur_h = _pc(line); cur_r = []; ic = False; hy = line[0]["top"]; continue
        if _is(lt):
            if cur_h and cur_r: tables.append({"h": cur_h, "r": cur_r, "c": ic})
            cur_h = None; cur_r = []; hy = None; ic = False; continue
        if cur_h is not None and hy is not None:
            if line[0]["top"] <= hy: continue
            row = _pp(_as(line, cur_h), cur_h)
            if sum(1 for c in row if c.strip()) >= 2: cur_r.append(row)
    if cur_h and cur_r: tables.append({"h": cur_h, "r": cur_r, "c": ic})
    return tables, cur_h if cur_h else None

# ─── Core conversion ────────────────────────────────────────────────
def convert_pdf_to_xlsx(pdf_stream):
    wb = Workbook(); wb.remove(wb.active); tc = 0; carry = None
    all_lists = {}; all_metrics = {}
    
    with pdfplumber.open(pdf_stream) as pdf:
        for pn, page in enumerate(pdf.pages, 1):
            # ── Widget/card extraction using plain text ──
            # Try full page text first
            full_text = page.extract_text(x_tolerance=3, y_tolerance=3)
            if full_text:
                l, m = _extract_widgets_from_text(full_text)
                for k, v in l.items():
                    if k not in all_lists or len(v) > len(all_lists[k]): all_lists[k] = v
                for k, v in m.items():
                    all_metrics[k] = v
            
            # Also try cropping left/right halves for side-by-side cards
            w = page.width; h = page.height
            for crop_box in [(0, 0, w * 0.52, h), (w * 0.48, 0, w, h)]:
                try:
                    cropped = page.crop(crop_box)
                    ct = cropped.extract_text(x_tolerance=3, y_tolerance=3)
                    if ct:
                        l, m = _extract_widgets_from_text(ct)
                        for k, v in l.items():
                            if k not in all_lists or len(v) > len(all_lists[k]): all_lists[k] = v
                        for k, v in m.items():
                            all_metrics[k] = v
                except: pass

            # ── Standard table extraction (original logic) ──
            tables, carry = _et(page, carry)
            for tbl in tables:
                hdrs, rows, ic = tbl["h"], tbl["r"], tbl["c"]
                keep = [(i, h) for i, h in enumerate(hdrs)
                        if h["name"].strip().lower() not in EXCLUDED_COLUMNS]
                if not keep or not rows: continue
                if ic and wb.sheetnames:
                    ws = wb[wb.sheetnames[-1]]; sr = ws.max_row + 1
                else:
                    tc += 1; ws = wb.create_sheet(title=_san(_strip_unicode(f"P{pn} Table {tc}")))
                    for oc, (_, h) in enumerate(keep, 1):
                        ws.cell(row=1, column=oc, value=_strip_unicode(h["name"]))
                    _sh(ws, 1, len(keep)); sr = 2
                for ri, row in enumerate(rows, sr):
                    for oc, (sc, h) in enumerate(keep, 1):
                        raw = row[sc] if sc < len(row) else ""
                        cn = h["name"].strip().lower()
                        if cn in NUMERIC_COLUMNS:
                            val = _efn(raw); isn = _in(val)
                        else: val = _strip_unicode(str(raw)) if raw else ""; isn = False
                        cell = ws.cell(row=ri, column=oc, value=val); _sc(cell, num=isn)
                _aw(ws)

    if all_metrics:
        ws = wb.create_sheet(title="Dashboard Metrics", index=0)
        for c, n in enumerate(["Metric", "Value", "Trend / Details"], 1): ws.cell(row=1, column=c, value=n)
        _sh(ws, 1, 3)
        for ri, (k, (v_num, v_det)) in enumerate(all_metrics.items(), 2):
            ws.cell(row=ri, column=1, value=_strip_unicode(k)); _sc(ws.cell(row=ri, column=1))
            val = _efn(v_num)
            c2 = ws.cell(row=ri, column=2, value=val); _sc(c2, num=_in(val))
            ws.cell(row=ri, column=3, value=_strip_unicode(v_det)); _sc(ws.cell(row=ri, column=3))
        _aw(ws)
        
    for list_name, rows in all_lists.items():
        if not rows: continue
        ws = wb.create_sheet(title=_san(_strip_unicode(list_name)), index=1)
        for c, n in enumerate(["Rank", "Name", "Volume"], 1): ws.cell(row=1, column=c, value=n)
        _sh(ws, 1, 3)
        for ri, r in enumerate(rows, 2):
            ws.cell(row=ri, column=1, value=int(r[0])); _sc(ws.cell(row=ri, column=1), num=True)
            ws.cell(row=ri, column=2, value=_strip_unicode(r[1])); _sc(ws.cell(row=ri, column=2))
            val = _efn(r[2])
            c3 = ws.cell(row=ri, column=3, value=val); _sc(c3, num=_in(val))
        _aw(ws)

    if tc == 0 and not all_metrics and not all_lists:
        ws = wb.create_sheet(title="Info", index=0)
        ws.cell(row=1, column=1, value="No tables found in this PDF.")
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return out

# ─── Streamlit UI ────────────────────────────────────────────────────
st.set_page_config(page_title="PDF → XLSX Converter", page_icon="📄", layout="centered")

st.markdown("""
<style>
.gradient-text {
    background: linear-gradient(135deg, #667eea, #764ba2);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    font-size: 2.8rem;
    font-weight: 700;
    margin-bottom: 0.5rem;
}
</style>
<div class="gradient-text">📄 WAF CONVERSION SOLUTION</div>
""", unsafe_allow_html=True)

st.markdown("Securely extract tables and convert security threat report PDFs to structured Excel spreadsheets.")

st.divider()

uploaded = st.file_uploader("Upload your PDF file", type=["pdf"],
                            help="Supports WAF threat report PDFs up to 50 MB")

if uploaded:
    sz = len(uploaded.getvalue())
    ss = f"{sz/1024:.1f} KB" if sz < 1_048_576 else f"{sz/1_048_576:.1f} MB"
    c1, c2 = st.columns(2)
    c1.info(f"📎 **{uploaded.name}**"); c2.info(f"📏 **{ss}**")

    if st.button("🔄 Convert to XLSX", use_container_width=True, type="primary"):
        with st.spinner("Extracting tables and converting..."):
            try:
                buf = convert_pdf_to_xlsx(uploaded)
                bn = uploaded.name.rsplit(".", 1)[0]
                st.success("✅ Conversion complete!")

                st.download_button("⬇️ Download XLSX", data=buf,
                    file_name=f"{bn}_converted.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True, type="primary")
            except Exception as e:
                st.error(f"❌ Conversion failed: {e}")

st.divider()
st.markdown("""
<div style='text-align: center; color: #888; font-size: 0.85rem; padding-bottom: 2rem;'>
  • Research & Intelligence • Proactive by design, intelligence with impact • Committed to ensuring Greatness
</div>
""", unsafe_allow_html=True)
